"""
Analytics Module 
Features:
- Super Admin: System-wide analytics and monitoring
- Group Manager: Group-specific analytics and insights  
- Participant: Personal activity dashboard
- Data visualization with charts and graphs
- CSV export functionality with preview
"""

from datetime import datetime, date, time, timedelta
from flask import render_template, request, redirect, url_for, flash, session, jsonify, make_response
from eventbridge_plus import app, db
from eventbridge_plus.auth import (
    require_login,
    require_platform_role,
    get_current_user_id,
    get_current_platform_role,
    is_super_admin,
    is_group_manager,
    get_user_home_url,
)
from eventbridge_plus.util import nz_date, nz_time12_upper
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font

# SUPER ADMIN ANALYTICS FUNCTIONS
def get_system_user_statistics(period='all'):
    """
    Get system-wide user statistics for Super Admin dashboard
    Corresponds to US-25: System-wide User Statistics
    
    Args:
        period: Time period filter - Note: User statistics always shows all users (users table has no created_at field)
    
    Returns:
        dict: User statistics including:
            - total_users: Total number of users in the system
            - role_breakdown: Count of users by role
            - status_breakdown: Count of users by status
            - recent_registrations: Users who created groups in last 30 days
    """
    try:
        # Note: users table doesn't have created_at field
        # Always show all users regardless of period parameter
        with db.get_cursor() as cursor:
            # Query 1: User counts and role/status breakdown
            cursor.execute("""
                SELECT
                COUNT(*) as total_users,
                SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) as active_users,
                SUM(CASE WHEN status = 'banned' THEN 1 ELSE 0 END) as banned_users,
                SUM(CASE WHEN platform_role = 'participant' THEN 1 ELSE 0 END) as participants,
                SUM(CASE WHEN platform_role = 'super_admin' THEN 1 ELSE 0 END) as super_admins,
                SUM(CASE WHEN platform_role = 'support_technician' THEN 1 ELSE 0 END) as support_techs
            FROM users
            """)
            user_stats = cursor.fetchone()
            
            # Validate that we got results
            if not user_stats or user_stats['total_users'] == 0:
                print("Warning: No users found in the system")
                return {
                    'total_users': 0,
                    'role_breakdown': {
                        'participants': 0,
                        'super_admins': 0,
                        'support_techs': 0
                    },
                    'status_breakdown': {
                        'active': 0,
                        'inactive': 0,
                        'banned': 0
                    },
                    'recent_registrations': 0
                }
            
            # Query 2: Recent registrations (last 30 days)
            # Uses group creation as a proxy for user activity
            cursor.execute("""
                SELECT COUNT(DISTINCT u.user_id) as recent_registrations
                FROM users u
                JOIN group_info g ON u.user_id = g.created_by
                WHERE g.created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
            """)
            recent_stats = cursor.fetchone()
            
            # Organize return data structure
            return {
                'total_users': user_stats['total_users'],
                'role_breakdown': {
                    'participants': user_stats['participants'] or 0,
                    'super_admins': user_stats['super_admins'] or 0,
                    'support_techs': user_stats['support_techs'] or 0
                },
                'status_breakdown': {
                    'active': user_stats['active_users'] or 0,
                    'banned': user_stats['banned_users'] or 0
                },
                'recent_registrations': recent_stats['recent_registrations'] or 0
            }
            
    except Exception as e:
        print(f"Error getting system user statistics: {e}")
        import traceback
        traceback.print_exc()
        return None


def get_event_participation_insights(period='all'):
    """
    Get event participation insights for Super Admin dashboard
    Corresponds to US-26: Event Participation Insights
    
    Args:
        period: Time period filter ('last_month', 'last_3_months', 'last_6_months', 'last_year', 'all')
    
    Returns:
        dict: Event statistics including:
            - event_stats: Counts by event status
            - participation_stats: Registration and attendance counts
            - rates: Calculated attendance and no-show rates
    """
    try:
        # Calculate date filter based on period
        from datetime import datetime, timedelta
        date_filter = ""
        date_params = []
        
        if period == 'last_month':
            start_date = datetime.now() - timedelta(days=30)
            date_filter = "WHERE event_date >= %s"
            date_params.append(start_date.strftime('%Y-%m-%d'))
        elif period == 'last_3_months':
            start_date = datetime.now() - timedelta(days=90)
            date_filter = "WHERE event_date >= %s"
            date_params.append(start_date.strftime('%Y-%m-%d'))
        elif period == 'last_6_months':
            start_date = datetime.now() - timedelta(days=180)
            date_filter = "WHERE event_date >= %s"
            date_params.append(start_date.strftime('%Y-%m-%d'))
        elif period == 'last_year':
            start_date = datetime.now() - timedelta(days=365)
            date_filter = "WHERE event_date >= %s"
            date_params.append(start_date.strftime('%Y-%m-%d'))
        
        with db.get_cursor() as cursor:
            # Query 1: Event statistics by status with date filter
            query = f"""
                SELECT 
                    COUNT(*) as total_events,
                    SUM(CASE WHEN status = 'scheduled' THEN 1 ELSE 0 END) as upcoming_events,
                    SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed_events,
                    SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_events
                FROM event_info
                {date_filter}
            """
            cursor.execute(query, tuple(date_params) if date_params else None)
            event_stats = cursor.fetchone()
            
            # ============================================
            # Query 2: Participation statistics with date filter
            # ============================================
            if date_filter:
                query = f"""
                    SELECT 
                        COUNT(*) as total_registrations,
                        SUM(CASE WHEN participation_status = 'attended' THEN 1 ELSE 0 END) as attended_count,
                        SUM(CASE WHEN participation_status = 'no_show' THEN 1 ELSE 0 END) as no_show_count,
                        SUM(CASE WHEN participation_status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_count
                    FROM event_members em
                    JOIN event_info e ON em.event_id = e.event_id
                    WHERE e.event_date >= %s
                """
                cursor.execute(query, (date_params[0],))
            else:
                cursor.execute("""
                    SELECT 
                        COUNT(*) as total_registrations,
                        SUM(CASE WHEN participation_status = 'attended' THEN 1 ELSE 0 END) as attended_count,
                        SUM(CASE WHEN participation_status = 'no_show' THEN 1 ELSE 0 END) as no_show_count,
                        SUM(CASE WHEN participation_status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_count
                    FROM event_members
                """)
            participation_stats = cursor.fetchone()
            # Calculate rates with safe division
            total_registrations = participation_stats['total_registrations'] or 0
            attended_count = participation_stats['attended_count'] or 0
            no_show_count = participation_stats['no_show_count'] or 0
            
            # Avoid division by zero
            attendance_rate = (attended_count / total_registrations * 100) if total_registrations > 0 else 0
            no_show_rate = (no_show_count / total_registrations * 100) if total_registrations > 0 else 0
            
            return {
                'event_stats': {
                    'total': event_stats['total_events'] or 0,
                    'upcoming': event_stats['upcoming_events'] or 0,
                    'completed': event_stats['completed_events'] or 0,
                    'cancelled': event_stats['cancelled_events'] or 0
                },
                'participation_stats': {
                    'total_registrations': total_registrations,
                    'attended': attended_count,
                    'no_show': no_show_count,
                    'cancelled': participation_stats['cancelled_count'] or 0
                },
                'rates': {
                    'attendance_rate': round(attendance_rate, 2),
                    'no_show_rate': round(no_show_rate, 2)
                }
            }
            
    except Exception as e:
        print(f"Error getting event participation insights: {e}")
        import traceback
        traceback.print_exc()
        return None


def get_platform_monitoring_data(period='all'):
    """
    Get platform-wide monitoring data for Super Admin dashboard
    
    Args:
        period: Time period filter - Note: Group Growth Trend always shows last 6 months for better visualization
    
    Note: The growth trend chart always shows last 6 months to ensure meaningful data points
    for trend visualization, regardless of the selected period filter.
    """


    try:
        with db.get_cursor() as cursor:
            # Query 1: Group statistics by status
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_groups,
                    SUM(CASE WHEN status = 'approved' THEN 1 ELSE 0 END) as active_groups,
                    SUM(CASE WHEN status = 'inactive' THEN 1 ELSE 0 END) as inactive_groups,
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending_groups
                FROM group_info
            """)
            group_stats = cursor.fetchone()
            

            # Calculate date range based on period and adjust data granularity
            from datetime import datetime, timedelta
            
            current_date = datetime.now()
            
            # Determine start date based on period
            if period == 'last_month':
                start_date = current_date - timedelta(days=30)
            elif period == 'last_3_months':
                start_date = current_date - timedelta(days=90)
            elif period == 'last_6_months':
                start_date = current_date - timedelta(days=180)
            elif period == 'last_year':
                start_date = current_date - timedelta(days=365)
            else:  # 'all'
                start_date = datetime(2020, 1, 1)
            
            # Format dates for SQL query
            start_date_str = start_date.strftime('%Y-%m-%d')
            end_date_str = current_date.strftime('%Y-%m-%d')
            
            # Always use monthly data for better visualization
            cursor.execute("""
                SELECT 
                    YEAR(created_at) as year,
                    MONTH(created_at) as month,
                    MIN(created_at) as first_date,
                    COUNT(*) as group_count
                FROM group_info
                WHERE status = 'approved'
                  AND created_at >= %s
                  AND created_at <= %s
                GROUP BY YEAR(created_at), MONTH(created_at)
                ORDER BY year ASC, month ASC
            """, (start_date_str, end_date_str))
            
            raw_data = cursor.fetchall()
            
            
            # Format dates using monthly format
            from datetime import datetime
            growth_trend = []
            
            for row in raw_data:
                first_date = row['first_date']
                
                # Handle different date formats
                if isinstance(first_date, str):
                    # If it's a string, parse it
                    try:
                        date_obj = datetime.strptime(first_date, '%Y/%m/%d %H:%M:%S')
                    except ValueError:
                        try:
                            date_obj = datetime.strptime(first_date, '%Y-%m-%d')
                        except ValueError:
                            print(f"Warning: Could not parse date: {first_date}")
                            continue
                elif isinstance(first_date, datetime):
                    # If it's already a datetime object
                    date_obj = first_date
                else:
                    # If it's a date object, convert to datetime
                    from datetime import date as date_type
                    if isinstance(first_date, date_type):
                        date_obj = datetime.combine(first_date, datetime.min.time())
                    else:
                        print(f"Warning: Unknown date type: {type(first_date)}")
                        continue
                
                # Format the date using NZ format
                from eventbridge_plus.util import nz_month_year
                month_label = nz_month_year(date_obj)
                
                # Add to growth trend list
                growth_trend.append({
                    'month': month_label,
                    'new_groups': row['group_count']
                })
            
            
            # Query 3: Activity summary
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT user_id) as active_users,
                    COALESCE(SUM(groups_joined), 0) as total_group_memberships,
                    COALESCE(SUM(events_participated), 0) as total_event_participations,
                    COALESCE(SUM(total_volunteer_hours), 0) as total_volunteer_hours
                FROM user_activity_summary
            """)
            activity_summary = cursor.fetchone()
            
            return {
                'group_stats': {
                    'total': group_stats['total_groups'] or 0,
                    'active': group_stats['active_groups'] or 0,
                    'inactive': group_stats['inactive_groups'] or 0,
                    'pending': group_stats['pending_groups'] or 0
                },
                'growth_trend': growth_trend,
                'activity_summary': {
                    'active_users': activity_summary['active_users'] or 0,
                    'total_group_memberships': activity_summary['total_group_memberships'] or 0,
                    'total_event_participations': activity_summary['total_event_participations'] or 0,
                    'total_volunteer_hours': activity_summary['total_volunteer_hours'] or 0
                }
            }
            
    except Exception as e:
        print(f" Error getting platform monitoring data: {e}")
        import traceback
        traceback.print_exc()
        return None


# EXPORT FUNCTIONS WITH NZ DATE FORMAT
def get_event_participation_data():
    """
    Get event participation data for preview and export
    Automatically loads data from the last 3 months based on current date
    All dates formatted in NZ format (DD/MM/YYYY)
    """
    try:
        # Calculate date range: from 3 months ago up to yesterday (exclude today)
        from datetime import datetime, timedelta
        
        current_date = datetime.now()
        three_months_ago = current_date - timedelta(days=90)  # Approximately 3 months
        yesterday = current_date - timedelta(days=1)
        
        # Format dates for SQL query
        start_date = three_months_ago.strftime('%Y-%m-%d')
        end_date = yesterday.strftime('%Y-%m-%d')
        
        with db.get_cursor() as cursor:
            cursor.execute("""
                SELECT 
                    e.event_id,
                    e.event_title,
                    e.event_type,
                    e.event_date,
                    e.event_time,
                    e.location,
                    e.status as event_status,
                    g.name as group_name,
                    g.group_type,
                    -- Only confirmed outcomes count toward Total
                    SUM(CASE WHEN em.participation_status IN ('attended','no_show','cancelled') THEN 1 ELSE 0 END) as total_registrations,
                    SUM(CASE WHEN em.participation_status = 'attended' THEN 1 ELSE 0 END) as attended,
                    SUM(CASE WHEN em.participation_status = 'no_show' THEN 1 ELSE 0 END) as no_show,
                    SUM(CASE WHEN em.participation_status = 'cancelled' THEN 1 ELSE 0 END) as cancelled,
                    SUM(CASE WHEN em.participation_status = 'registered' THEN 1 ELSE 0 END) as registered,
                    ROUND(
                        CASE 
                            WHEN SUM(CASE WHEN em.participation_status IN ('attended','no_show','cancelled') THEN 1 ELSE 0 END) > 0 
                            THEN SUM(CASE WHEN em.participation_status = 'attended' THEN 1 ELSE 0 END) * 100.0 / 
                                 SUM(CASE WHEN em.participation_status IN ('attended','no_show','cancelled') THEN 1 ELSE 0 END)
                            ELSE 0 
                        END, 2
                    ) as attendance_rate
                FROM event_info e
                JOIN group_info g ON e.group_id = g.group_id
                LEFT JOIN event_members em ON e.event_id = em.event_id
                WHERE e.status = 'completed'
                  AND e.event_date >= %s
                  AND e.event_date <= %s
                GROUP BY e.event_id, e.event_title, e.event_type, e.event_date, 
                         e.event_time, e.location, e.status, g.name, g.group_type
                ORDER BY e.event_date DESC, e.event_id DESC
            """, (start_date, end_date))
            data = cursor.fetchall()
            
            # Format dates and times using NZ format
            for row in data:
                row['formatted_date'] = nz_date(row['event_date'])      # DD/MM/YYYY format
                row['formatted_time'] = nz_time12_upper(row['event_time'])  # 12-hour format
            
            return data
            
    except Exception as e:
        print(f"Error getting event participation data: {e}")
        import traceback
        traceback.print_exc()
        return []


def export_event_participation_csv():
    """
    Export event participation data as CSV with NZ date format
    All dates formatted as DD/MM/YYYY
    
    Returns:
        Response: CSV file download response with UTF-8 encoding
    """
    try:
        data = get_event_participation_data()
        
        if not data:
            return None
        
        # Create CSV using StringIO
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header with descriptive column names
        writer.writerow([
            'Event ID',
            'Event Title',
            'Event Type',
            'Date (DD/MM/YYYY)',  # Explicitly show NZ format
            'Time',
            'Location',
            'Status',
            'Group Name',
            'Group Type',
            'Total Registrations',
            'Attended',
            'No Show',
            'Cancelled',
            'Registered',
            'Attendance Rate (%)'
        ])
        
        # Write data rows with NZ date format
        for row in data:
            writer.writerow([
                row['event_id'],
                row['event_title'],
                row['event_type'],
                row['formatted_date'],  # NZ date format (DD/MM/YYYY)
                row['formatted_time'],  # 12-hour format with AM/PM
                row['location'] or 'N/A',
                row['event_status'].title(),
                row['group_name'],
                row['group_type'].title(),
                row['total_registrations'],
                row['attended'],
                row['no_show'],
                row['cancelled'],
                row['registered'],
                row['attendance_rate']
            ])
        
        # Create response with timestamp
        csv_content = output.getvalue()
        output.close()
        
        timestamp = datetime.now().strftime("%Y%m%d")
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'  # UTF-8 with BOM for Excel
        response.headers['Content-Disposition'] = f'attachment; filename=event_participation_report_{timestamp}.csv'
        
        return response
        
    except Exception as e:
        print(f"Error exporting event participation CSV: {e}")
        import traceback
        traceback.print_exc()
        return None


def export_event_participation_xlsx():
    """
    Export event data as Excel file
    """
    try:
        data = get_event_participation_data()
        if not data:
            return None
        
        # Create excel file
        wb = Workbook()
        ws = wb.active
        
        # Write headers
        headers = [
            'Event ID', 'Event Title', 'Event Type', 'Date', 'Time',
            'Location', 'Status', 'Group Name', 'Group Type',
            'Total Registrations', 'Attended', 'No Show', 'Cancelled',
            'Registered', 'Attendance Rate (%)'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)  # Make header bold
        
        # Write data
        for idx, row in enumerate(data, 2):
            ws.cell(row=idx, column=1).value = row['event_id']
            ws.cell(row=idx, column=2).value = row['event_title']
            ws.cell(row=idx, column=3).value = row['event_type']
            ws.cell(row=idx, column=4).value = row['formatted_date']
            ws.cell(row=idx, column=5).value = row['formatted_time']
            ws.cell(row=idx, column=6).value = row['location'] or 'N/A'
            ws.cell(row=idx, column=7).value = row['event_status'].title()
            ws.cell(row=idx, column=8).value = row['group_name']
            ws.cell(row=idx, column=9).value = row['group_type'].title()
            ws.cell(row=idx, column=10).value = row['total_registrations']
            ws.cell(row=idx, column=11).value = row['attended']
            ws.cell(row=idx, column=12).value = row['no_show']
            ws.cell(row=idx, column=13).value = row['cancelled']
            ws.cell(row=idx, column=14).value = row['registered']
            ws.cell(row=idx, column=15).value = row['attendance_rate']
        
        # Save file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file for download
        timestamp = datetime.now().strftime("%Y%m%d")
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=event_report_{timestamp}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error creating xlsx: {e}")
        return None


# GROUP MANAGER ANALYTICS FUNCTIONS
def get_group_analytics(group_id, activity_period='last_6_months', attendance_period='last_6_months'):
    """
    Get comprehensive analytics for a specific group (for Group Manager)
    Includes member statistics, activity data, etc.
    
    Args:
        group_id: Group ID to analyze
        activity_period: Time period filter for activity levels ('last_month', 'last_3_months', 'last_6_months', 'last_year', 'all')
        attendance_period: Time period filter for attendance trends ('last_3_months', 'last_6_months', 'last_year')
    """
    try:
        from eventbridge_plus.util import nz_month_year
        from datetime import datetime, timedelta
        
        current_date = datetime.now()
        
        # Calculate date range for activity period
        def get_period_dates(period):
            if period == 'last_month':
                return current_date - timedelta(days=30), 1
            elif period == 'last_3_months':
                return current_date - timedelta(days=90), 3
            elif period == 'last_6_months':
                return current_date - timedelta(days=180), 6
            elif period == 'last_year':
                return current_date - timedelta(days=365), 12
            else:  # 'all'
                return datetime(2020, 1, 1), 100
        
        start_date, months_back = get_period_dates(activity_period)
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = current_date.strftime('%Y-%m-%d')
        
        # Calculate date range for attendance period
        attendance_start_date, attendance_months_back = get_period_dates(attendance_period)
        attendance_start_date_str = attendance_start_date.strftime('%Y-%m-%d')
        attendance_end_date_str = current_date.strftime('%Y-%m-%d')
        
        with db.get_cursor() as cursor:
            # Get group info and manager name
            cursor.execute("""
                SELECT g.group_id, g.name, g.group_type, g.status, g.created_at,
                       u.username as manager_username, u.first_name, u.last_name
                FROM group_info g
                LEFT JOIN group_members gm ON g.group_id = gm.group_id AND gm.group_role = 'manager'
                LEFT JOIN users u ON gm.user_id = u.user_id
                WHERE g.group_id = %s
            """, (group_id,))
            
            group_info = cursor.fetchone()
            
            if not group_info:
                return None
            
            # Member Statistics 
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_members,
                    SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) as active_members,
                    SUM(CASE WHEN status = 'left' THEN 1 ELSE 0 END) as left_members,
                    SUM(CASE WHEN group_role = 'manager' THEN 1 ELSE 0 END) as managers,
                    SUM(CASE WHEN group_role = 'volunteer' THEN 1 ELSE 0 END) as volunteers,
                    SUM(CASE WHEN group_role = 'member' THEN 1 ELSE 0 END) as members
                FROM group_members
                WHERE group_id = %s
            """, (group_id,))
            
            member_stats = cursor.fetchone()
            
            # Calculate role distribution percentages for pie chart
            managers_count = int(member_stats['managers'] or 0)
            volunteers_count = int(member_stats['volunteers'] or 0)
            members_count = int(member_stats['members'] or 0)
            active_total = managers_count + volunteers_count + members_count
            
            if active_total > 0:
                role_distribution = [
                    {
                        "role": "Managers",
                        "count": managers_count,
                        "percentage": round(managers_count * 100.0 / active_total, 1)
                    },
                    {
                        "role": "Volunteers", 
                        "count": volunteers_count,
                        "percentage": round(volunteers_count * 100.0 / active_total, 1)
                    },
                    {
                        "role": "Members",
                        "count": members_count,
                        "percentage": round(members_count * 100.0 / active_total, 1)
                    }
                ]
            else:
                role_distribution = []
            
            # Event Statistics 
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_events,
                    SUM(CASE WHEN status = 'scheduled' THEN 1 ELSE 0 END) as scheduled_events,
                    SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed_events,
                    SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_events
                FROM event_info
                WHERE group_id = %s
            """, (group_id,))
            
            event_stats = cursor.fetchone()
            
            # Volunteer Statistics 
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT u.user_id) as total_volunteers,
                    COALESCE(SUM(em.volunteer_hours), 0) as total_volunteer_hours,
                    COALESCE(AVG(em.volunteer_hours), 0) as avg_volunteer_hours
                FROM users u
                JOIN group_members gm ON u.user_id = gm.user_id
                LEFT JOIN event_members em ON u.user_id = em.user_id AND em.event_role = 'volunteer'
                LEFT JOIN event_info e ON em.event_id = e.event_id AND e.group_id = %s
                WHERE gm.group_id = %s AND gm.status = 'active' AND em.volunteer_hours > 0
            """, (group_id, group_id))
            
            volunteer_stats = cursor.fetchone()
            
            # New Members (Last 30 Days)
            cursor.execute("""
                SELECT COUNT(*) as new_members_30_days
                FROM group_members
                WHERE group_id = %s 
                AND join_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                AND status = 'active'
            """, (group_id,))
            
            new_members_stats = cursor.fetchone()
            
            # Average Participation Rate
            cursor.execute("""
                SELECT 
                    ROUND(AVG(participation_rate), 1) as avg_participation_rate
                FROM (
                    SELECT 
                        CASE 
                            WHEN COUNT(em.membership_id) > 0 
                            THEN SUM(CASE WHEN em.participation_status = 'attended' THEN 1 ELSE 0 END) * 100.0 / COUNT(em.membership_id)
                            ELSE 0 
                        END as participation_rate
                    FROM event_info e
                    LEFT JOIN event_members em ON e.event_id = em.event_id
                    WHERE e.group_id = %s AND e.status = 'completed'
                    GROUP BY e.event_id
                ) as event_rates
            """, (group_id,))
            
            avg_participation_result = cursor.fetchone()
            avg_participation_rate = avg_participation_result['avg_participation_rate'] if avg_participation_result and avg_participation_result['avg_participation_rate'] is not None else 0
            
            # Total Volunteer Hours (Last 6 Months)
            # Ensure start date is not before group creation date
            group_created_at = group_info['created_at']
            if isinstance(group_created_at, str):
                group_created_at = datetime.strptime(group_created_at, '%Y-%m-%d %H:%M:%S')
            elif isinstance(group_created_at, date) and not isinstance(group_created_at, datetime):
                group_created_at = datetime.combine(group_created_at, datetime.min.time())
            
            # Use the later of calculated start_date or group creation date
            actual_start_date = max(start_date, group_created_at)
            actual_start_date_str = actual_start_date.strftime('%Y-%m-%d')
            
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(em.volunteer_hours), 0) as total_volunteer_hours_6m
                FROM event_members em
                JOIN event_info e ON em.event_id = e.event_id
                WHERE e.group_id = %s 
                AND e.event_date >= %s
                AND e.event_date <= %s
                AND em.event_role = 'volunteer'
                AND em.volunteer_hours IS NOT NULL
            """, (group_id, actual_start_date_str, end_date_str))
            
            volunteer_hours_6m_stats = cursor.fetchone()
            
            # Per-Event Participation Rates (only past and today events)
            # Rate % calculation: (Attended / Total Participants) * 100
            # Total Participants = Attended + No Show + Cancelled (excluding registered status)
            # This represents the attendance rate among participants with confirmed status
            cursor.execute("""
                SELECT 
                    e.event_id,
                    e.event_title,
                    e.event_date,
                    SUM(CASE WHEN em.participation_status IN ('attended', 'no_show', 'cancelled') THEN 1 ELSE 0 END) as total_participants,
                    SUM(CASE WHEN em.participation_status = 'attended' THEN 1 ELSE 0 END) as attended_count,
                    SUM(CASE WHEN em.participation_status = 'no_show' THEN 1 ELSE 0 END) as no_show_count,
                    SUM(CASE WHEN em.participation_status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_count,
                    ROUND(
                        CASE 
                            WHEN SUM(CASE WHEN em.participation_status IN ('attended', 'no_show', 'cancelled') THEN 1 ELSE 0 END) > 0 
                            THEN SUM(CASE WHEN em.participation_status = 'attended' THEN 1 ELSE 0 END) * 100.0 / SUM(CASE WHEN em.participation_status IN ('attended', 'no_show', 'cancelled') THEN 1 ELSE 0 END)
                            ELSE 0 
                        END, 1
                    ) as participation_rate
                FROM event_info e
                LEFT JOIN event_members em ON e.event_id = em.event_id
                WHERE e.group_id = %s AND e.event_date <= CURDATE()
                GROUP BY e.event_id, e.event_title, e.event_date
                ORDER BY e.event_date DESC
            """, (group_id,))
            
            participation_rates = cursor.fetchall()
            
            # Top 5 Participants - Find participants with most events attended
            cursor.execute("""
                SELECT 
                    u.user_id,
                    u.username,
                    CONCAT(u.first_name, ' ', u.last_name) as full_name,
                    COUNT(DISTINCT CASE WHEN em.participation_status = 'attended' THEN em.event_id END) as events_attended,
                    COALESCE(SUM(em.volunteer_hours), 0) as volunteer_hours
                FROM users u
                JOIN group_members gm ON u.user_id = gm.user_id
                LEFT JOIN event_members em ON u.user_id = em.user_id
                LEFT JOIN event_info e ON em.event_id = e.event_id AND e.group_id = %s
                WHERE gm.group_id = %s AND gm.status = 'active'
                GROUP BY u.user_id, u.username, u.first_name, u.last_name
                ORDER BY events_attended DESC
                LIMIT 5
            """, (group_id, group_id))
            
            top_participants = cursor.fetchall()
            
            # Top 5 Volunteers - Find volunteers with most hours (last 6 months)
            # Use actual_start_date_str to ensure dates are not before group creation
            cursor.execute("""
                SELECT 
                    u.user_id,
                    u.username,
                    CONCAT(u.first_name, ' ', u.last_name) as full_name,
                    COALESCE(SUM(em.volunteer_hours), 0) as volunteer_hours,
                    COUNT(DISTINCT em.event_id) as volunteer_events
                FROM users u
                JOIN group_members gm ON u.user_id = gm.user_id
                JOIN event_members em ON u.user_id = em.user_id AND em.event_role = 'volunteer'
                JOIN event_info e ON em.event_id = e.event_id 
                    AND e.group_id = %s 
                    AND e.event_date >= %s
                    AND e.event_date <= %s
                WHERE gm.group_id = %s AND gm.status = 'active'
                AND em.volunteer_hours IS NOT NULL
                GROUP BY u.user_id, u.username, u.first_name, u.last_name
                ORDER BY volunteer_hours DESC
                LIMIT 5
            """, (group_id, actual_start_date_str, end_date_str, group_id))
            
            top_volunteers = cursor.fetchall()
            
            # Calculate member activity levels - based on participation rate (last 6 months dynamically)
            # Use actual_start_date_str to ensure dates are not before group creation
            cursor.execute("""
                WITH member_participation AS (
                    SELECT 
                        gm.user_id,
                        COUNT(DISTINCT ei.event_id) as total_group_events,
                        COUNT(DISTINCT em.event_id) as attended_events,
                        CASE 
                            WHEN COUNT(DISTINCT ei.event_id) = 0 THEN 0
                            ELSE ROUND(COUNT(DISTINCT em.event_id) * 100.0 / COUNT(DISTINCT ei.event_id), 1)
                        END as participation_rate
                    FROM group_members gm
                    LEFT JOIN event_info ei ON ei.group_id = gm.group_id 
                        AND ei.status = 'completed'
                        AND ei.event_date >= %s
                        AND ei.event_date <= %s
                    LEFT JOIN event_members em ON em.event_id = ei.event_id 
                        AND em.user_id = gm.user_id
                        AND em.participation_status = 'attended'
                    WHERE gm.group_id = %s
                      AND gm.status = 'active'
                    GROUP BY gm.user_id
                )
                SELECT 
                    SUM(CASE WHEN participation_rate >= 80 THEN 1 ELSE 0 END) as very_active,
                    SUM(CASE WHEN participation_rate >= 50 AND participation_rate < 80 THEN 1 ELSE 0 END) as active,
                    SUM(CASE WHEN participation_rate >= 20 AND participation_rate < 50 THEN 1 ELSE 0 END) as moderate,
                    SUM(CASE WHEN participation_rate < 20 THEN 1 ELSE 0 END) as low_activity
                FROM member_participation
            """, (actual_start_date_str, end_date_str, group_id))
            
            activity_row = cursor.fetchone()
            activity_breakdown = {
                'very_active': activity_row['very_active'] or 0,
                'active': activity_row['active'] or 0,
                'moderate': activity_row['moderate'] or 0,
                'low_activity': activity_row['low_activity'] or 0
            }
            
            # Engagement Trends - Generate complete series based on period
            import calendar
            
            # Generate list of months for attendance trend (use attendance_period)
            attendance_months = []
            attendance_start_month = current_date.replace(day=1)
            
            # Go back the appropriate number of months for attendance
            for i in range(attendance_months_back):
                if i == 0:
                    month_date = attendance_start_month
                else:
                    # Go back one month properly
                    if month_date.month == 1:
                        month_date = month_date.replace(year=month_date.year - 1, month=12)
                    else:
                        month_date = month_date.replace(month=month_date.month - 1)
                attendance_months.append(month_date)
            
            attendance_months.reverse()  # Order from oldest to newest
            
            # Get actual attendance data - include both 'attended' and 'registered' for past events
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(e.event_date, '%%Y-%%m-01') as month_date,
                    COUNT(DISTINCT CASE 
                        WHEN em.participation_status = 'attended' THEN em.user_id 
                        WHEN em.participation_status = 'registered' AND e.event_date < CURDATE() THEN em.user_id
                        ELSE NULL 
                    END) as attended_count
                FROM event_info e
                LEFT JOIN event_members em ON e.event_id = em.event_id
                WHERE e.group_id = %s 
                    AND e.event_date >= %s
                    AND e.event_date <= %s
                    AND e.status IN ('completed', 'scheduled', 'ongoing')
                GROUP BY DATE_FORMAT(e.event_date, '%%Y-%%m-01')
                ORDER BY month_date ASC
            """, (group_id, attendance_start_date_str, attendance_end_date_str))
            
            raw_trends = cursor.fetchall()
            
            # Create a dictionary for quick lookup
            trends_dict = {}
            for row in raw_trends:
                month_date = row['month_date']
                if isinstance(month_date, str):
                    month_date = datetime.strptime(month_date, '%Y-%m-%d')
                trends_dict[month_date.strftime('%Y-%m-01')] = row['attended_count']
            
            # Format dates for engagement trends - ensure all months are included
            engagement_trends = []
            for month_date in attendance_months:
                month_key = month_date.strftime('%Y-%m-01')
                attended_count = trends_dict.get(month_key, 0)
                engagement_trends.append({
                    'month': nz_month_year(month_date),
                    'attended_count': attended_count
                })
            
            # Convert Decimal types to int for JSON serialization
            member_stats = convert_decimals_to_int(member_stats)
            role_distribution = convert_decimals_to_int(role_distribution)
            event_stats = convert_decimals_to_int(event_stats)
            volunteer_stats = convert_decimals_to_int(volunteer_stats)
            new_members_stats = convert_decimals_to_int(new_members_stats)
            volunteer_hours_6m_stats = convert_decimals_to_int(volunteer_hours_6m_stats)
            participation_rates = convert_decimals_to_int(participation_rates)
            top_participants = convert_decimals_to_int(top_participants)
            top_volunteers = convert_decimals_to_int(top_volunteers)
            activity_breakdown = convert_decimals_to_int(activity_breakdown)
            
            # Format date range for display (use actual_start_date for volunteer hours display)
            from eventbridge_plus.util import nz_date
            # Use actual_start_date that considers group creation date
            start_date_obj = actual_start_date
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d')
            date_range_display = f"{nz_date(start_date_obj)} - {nz_date(end_date_obj)}"
            
            return {
                'group_info': group_info,
                'member_stats': member_stats,
                'role_distribution': role_distribution,
                'event_stats': event_stats,
                'volunteer_stats': volunteer_stats,
                'new_members_stats': new_members_stats,
                'avg_participation_rate': avg_participation_rate,
                'volunteer_hours_6m_stats': volunteer_hours_6m_stats,
                'participation_rates': participation_rates,
                'top_participants': top_participants,
                'top_volunteers': top_volunteers,
                'activity_breakdown': activity_breakdown,
                'engagement_trends': engagement_trends,
                'date_range_display': date_range_display
            }
            
    except Exception as e:
        print(f"Error getting group analytics for group {group_id}: {e}")
        import traceback
        traceback.print_exc()
        return None


# ANALYTICS ROUTES
@app.route('/analytics/super-admin')
@require_login
@require_platform_role('super_admin')
def super_admin_analytics():
    """
    Super Admin analytics dashboard route
    """
    try:
        # Get all analytics data
        user_stats = get_system_user_statistics()
        event_insights = get_event_participation_insights()
        platform_data = get_platform_monitoring_data()
        
        # Check if any data failed to load
        if user_stats is None:
            flash('Unable to load user statistics. Please try again.', 'warning')
        if event_insights is None:
            flash('Unable to load event insights. Please try again.', 'warning')
        if platform_data is None:
            flash('Unable to load platform data. Please try again.', 'warning')
        
        return render_template('admin_home.html',
                             user_stats=user_stats,
                             event_insights=event_insights,
                             platform_data=platform_data)
                             
    except Exception as e:
        print(f"Error loading super admin analytics: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading analytics dashboard. Please try again.', 'error')
        return redirect(url_for('admin_dashboard'))


# EXPORT ROUTES
@app.route('/analytics/export/events/preview')
@require_login
@require_platform_role('super_admin')
def preview_events_export():
    """
    Preview event participation data before downloading
    Shows a table view of all data that will be exported
    All dates displayed in NZ format (DD/MM/YYYY)
    """
    try:
        from eventbridge_plus.util import get_pagination_params, create_pagination_info, create_pagination_links
        
        # Get pagination parameters
        page, per_page = get_pagination_params(request, default_per_page=20)
        
        # Get all data first
        all_data = get_event_participation_data()
        
        if not all_data:
            flash('No event data available for export.', 'warning')
            return redirect(url_for('admin_dashboard'))
        
        # Apply pagination
        total_records = len(all_data)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        data = all_data[start_idx:end_idx]
        
        # Create pagination info
        base_url = url_for('preview_events_export')
        pagination = create_pagination_info(
            page=page,
            per_page=per_page,
            total=total_records,
            base_url=base_url
        )
        pagination_links = create_pagination_links(pagination)
        
        # Build period range string (last 3 months up to yesterday)
        from datetime import datetime, timedelta
        now_dt = datetime.now()
        start_dt = now_dt - timedelta(days=90)
        end_dt = now_dt - timedelta(days=1)
        period_range = f"{nz_date(start_dt)} - {nz_date(end_dt)}"
        
        return render_template('analytics/export_preview.html',
                             data=data,
                             total_records=total_records,
                             period_range=period_range,
                             pagination=pagination,
                             pagination_links=pagination_links)
                             
    except Exception as e:
        print(f"Error loading export preview: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading export preview. Please try again.', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/analytics/export/events/download')
@require_login
@require_platform_role('super_admin')
def download_events_csv():
    """
    Download event participation data as CSV
    All dates formatted in NZ format (DD/MM/YYYY)
    
    This is the actual download endpoint
    """
    try:
        csv_response = export_event_participation_csv()
        
        if csv_response:
            flash('Event data exported successfully!', 'success')
            return csv_response
        else:
            flash('Error generating CSV export. Please try again.', 'error')
            return redirect(url_for('preview_events_export'))
            
    except Exception as e:
        print(f"Error downloading events CSV: {e}")
        import traceback
        traceback.print_exc()
        flash('Error exporting data. Please try again.', 'error')
        return redirect(url_for('preview_events_export'))


@app.route('/analytics/export/events/download-xlsx')
@require_login
@require_platform_role('super_admin')
def download_events_xlsx():
    """
    Download event data as Excel file
    """
    try:
        xlsx_response = export_event_participation_xlsx()
        
        if xlsx_response:
            return xlsx_response
        else:
            flash('Error creating Excel file', 'error')
            return redirect(url_for('preview_events_export'))
            
    except Exception as e:
        print(f"Error: {e}")
        flash('Error exporting data', 'error')
        return redirect(url_for('preview_events_export'))


@app.route('/analytics/export/events')
@require_login
@require_platform_role('super_admin')
def export_events_csv():
    """
    Redirect to preview page instead of direct download
    (Keeping old route for backward compatibility)
    """
    return redirect(url_for('preview_events_export'))


@app.route('/analytics/group/<int:group_id>')
@require_login
def group_analytics(group_id):
    """
    Group-specific analytics dashboard (for Group Managers)
    Corresponds to US-28: Group-specific Analytics
    
    Args:
        group_id: The group to display analytics for
    """
    try:
        print(f"Loading analytics for group_id: {group_id}")
        
        # Check if user is super admin or support technician (can access all groups)
        if is_super_admin() or get_current_platform_role() == 'support_technician':
            print("User is super admin or support technician - access granted")
            # Super admin and support technician can access any group analytics
            pass
        else:
            # Verify user is a manager of this group
            user_id = get_current_user_id()
            print(f"Checking permissions for user_id: {user_id}")
            
            with db.get_cursor() as cursor:
                cursor.execute("""
                    SELECT group_role 
                    FROM group_members 
                    WHERE group_id = %s AND user_id = %s AND status = 'active'
                """, (group_id, user_id))
                
                membership = cursor.fetchone()
                print(f"Membership result: {membership}")
                
                if not membership:
                    print("No membership found - redirecting")
                    flash('You do not have access to this group.', 'error')
                    return redirect(get_user_home_url())
                
                if membership['group_role'] != 'manager':
                    print(f"User role is {membership['group_role']}, not manager - redirecting")
                    flash('Only group managers can view analytics.', 'error')
                    return redirect(url_for('group_detail', group_id=group_id))
                else:
                    print("User is manager - access granted")
        
        # Get period filters from request (two independent filters)
        from flask import request
        activity_period = request.args.get('activity_period', 'last_3_months')
        attendance_period = request.args.get('attendance_period', 'last_6_months')
        
        # Get analytics data with both period parameters
        analytics_data = get_group_analytics(group_id, 
                                           activity_period=activity_period,
                                           attendance_period=attendance_period)
        
        if not analytics_data:
            flash('Unable to load group analytics. Please check if the group exists and you have access.', 'error')
            return redirect(url_for('group_detail', group_id=group_id))
        
        # Map periods to display labels
        period_labels = {
            'last_month': 'Last Month',
            'last_3_months': 'Last 3 Months',
            'last_6_months': 'Last 6 Months',
            'last_year': 'Last Year',
            'all': 'All Time'
        }
        activity_period_label = period_labels.get(activity_period, 'Last 6 Months')
        attendance_period_label = period_labels.get(attendance_period, 'Last 6 Months')
        
        return render_template('analytics/group_analytics.html',
                             analytics=analytics_data,
                             activity_period=activity_period,
                             attendance_period=attendance_period,
                             activity_period_label=activity_period_label,
                             attendance_period_label=attendance_period_label)
                             
    except Exception as e:
        print(f"Error loading group analytics: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading group analytics. Please try again.', 'error')
        return redirect(get_user_home_url())


# =============================================================================
# SUPPORT TECHNICIAN ANALYTICS FUNCTIONS
# =============================================================================

@app.route('/analytics/helpdesk')
@app.route('/helpdesk-analytics')
@require_login
@require_platform_role('support_technician', 'super_admin')
def helpdesk_analytics():
    """
    Helpdesk Analytics Dashboard
    """
    try:
        # Get period filter from request
        period = request.args.get('period', 'all')
        
        # Get helpdesk statistics
        helpdesk_stats = get_helpdesk_statistics(period)
        
        # Get support performance metrics
        performance_metrics = get_support_performance_metrics(period)
        
        # Get request trends (past 6 months)
        request_trends = get_request_trends(period)
        
        # Get category breakdown
        category_breakdown = get_category_breakdown(period)
        
        return render_template('analytics/helpdesk_analytics.html',
                             helpdesk_stats=helpdesk_stats,
                             performance_metrics=performance_metrics,
                             request_trends=request_trends,
                             category_breakdown=category_breakdown,
                             current_period=period)
                             
    except Exception as e:
        print(f"Error loading helpdesk analytics: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading helpdesk analytics. Please try again.', 'error')
        return redirect(url_for('support_dashboard'))



def get_helpdesk_statistics(period='all'):
    """
    Get helpdesk statistics for Support Technician analytics
    Corresponds to US-29: Support Technician Analytics Dashboard
    """
    try:
        with db.get_cursor() as cursor:
            # Build date filter based on period
            date_filter = ""
            if period == 'last_week':
                date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 1 WEEK)"
            elif period == 'last_month':
                date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 1 MONTH)"
            elif period == 'last_3_months':
                date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 3 MONTH)"
            elif period == 'last_6_months':
                date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 6 MONTH)"
            
            # Get basic statistics
            cursor.execute(f"""
                SELECT
                    COUNT(*) as total_requests,
                    SUM(CASE WHEN status = 'new' THEN 1 ELSE 0 END) as new_requests,
                    SUM(CASE WHEN status = 'assigned' THEN 1 ELSE 0 END) as assigned_requests,
                    SUM(CASE WHEN status = 'solved' THEN 1 ELSE 0 END) as solved_requests,
                    SUM(CASE WHEN status = 'blocked' THEN 1 ELSE 0 END) as blocked_requests
                FROM help_requests
                WHERE 1=1 {date_filter}
            """)
            stats = cursor.fetchone()
            
            # Get resolved in period (or this month if all time)
            if period == 'all':
                resolved_query = """
                    SELECT COUNT(*) as resolved_this_month
                    FROM help_requests
                    WHERE status = 'solved' 
                    AND MONTH(updated_at) = MONTH(CURRENT_DATE())
                    AND YEAR(updated_at) = YEAR(CURRENT_DATE())
                """
            else:
                resolved_query = f"""
                    SELECT COUNT(*) as resolved_this_month
                    FROM help_requests
                    WHERE status = 'solved' {date_filter}
                """
            
            cursor.execute(resolved_query)
            monthly_stats = cursor.fetchone()
            
            # Get active support staff count
            cursor.execute("""
                SELECT COUNT(*) as active_support_staff
                FROM users
                WHERE platform_role IN ('support_technician', 'super_admin')
                AND status = 'active'
            """)
            staff_stats = cursor.fetchone()
            
            total = stats['total_requests'] or 0
            new_count = stats['new_requests'] or 0
            assigned_count = stats['assigned_requests'] or 0
            solved_count = stats['solved_requests'] or 0
            blocked_count = stats['blocked_requests'] or 0
            
            return convert_decimals_to_int({
                'total_requests': total,
                'new_requests': new_count,
                'assigned_requests': assigned_count,
                'solved_requests': solved_count,
                'blocked_requests': blocked_count,
                'resolved_this_month': monthly_stats['resolved_this_month'] or 0,
                'active_support_staff': staff_stats['active_support_staff'] or 0,
                # Add percentage calculations
                'new_percentage': round((new_count / total * 100), 1) if total > 0 else 0,
                'assigned_percentage': round((assigned_count / total * 100), 1) if total > 0 else 0,
                'solved_percentage': round((solved_count / total * 100), 1) if total > 0 else 0,
                'blocked_percentage': round((blocked_count / total * 100), 1) if total > 0 else 0
            })
            
    except Exception as e:
        print(f"Error getting helpdesk statistics: {e}")
        return {
            'total_requests': 0,
            'new_requests': 0,
            'assigned_requests': 0,
            'solved_requests': 0,
            'blocked_requests': 0,
            'resolved_this_month': 0,
            'active_support_staff': 0
        }


def get_support_performance_metrics(period='all'):
    """
    Get support performance metrics
    """
    try:
        with db.get_cursor() as cursor:
            # Build date filter based on period
            date_filter = ""
            if period == 'last_week':
                date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 1 WEEK)"
            elif period == 'last_month':
                date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 1 MONTH)"
            elif period == 'last_3_months':
                date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 3 MONTH)"
            elif period == 'last_6_months':
                date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 6 MONTH)"
            
            # Get average response time (in hours)
            cursor.execute(f"""
                SELECT AVG(TIMESTAMPDIFF(HOUR, created_at, last_staff_reply_at)) as avg_response_time
                FROM help_requests
                WHERE last_staff_reply_at IS NOT NULL {date_filter}
            """)
            response_time = cursor.fetchone()
            
            # Get resolution rate
            cursor.execute(f"""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN status = 'solved' THEN 1 ELSE 0 END) as resolved
                FROM help_requests
                WHERE 1=1 {date_filter}
            """)
            resolution_stats = cursor.fetchone()
            
            resolution_rate = 0
            if resolution_stats['total'] > 0:
                resolution_rate = (resolution_stats['resolved'] / resolution_stats['total']) * 100
            
            return convert_decimals_to_int({
                'avg_response_time': response_time['avg_response_time'] or 0,
                'resolution_rate': round(resolution_rate, 1)
            })
            
    except Exception as e:
        print(f"Error getting support performance metrics: {e}")
        return {
            'avg_response_time': 0,
            'resolution_rate': 0
        }


def get_request_trends(period='all'):
    """
    Get request trends for the specified period
    """
    try:
        from eventbridge_plus.util import nz_date
        from datetime import datetime
        
        with db.get_cursor() as cursor:
            # Build date filter based on period
            date_filter = ""
            if period == 'last_week':
                date_filter = "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 1 WEEK)"
            elif period == 'last_month':
                date_filter = "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 1 MONTH)"
            elif period == 'last_3_months':
                date_filter = "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 3 MONTH)"
            elif period == 'last_6_months':
                date_filter = "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 6 MONTH)"
            
            # For shorter periods, use daily data instead of monthly
            if period in ['last_week', 'last_month']:
                cursor.execute(f"""
                    SELECT 
                        DATE(created_at) as date,
                        COUNT(*) as request_count
                    FROM help_requests
                    {date_filter}
                    GROUP BY DATE(created_at)
                    ORDER BY date ASC
                """)
            else:
                cursor.execute(f"""
                    SELECT 
                        DATE_FORMAT(created_at, '%Y-%m-01') as month_date,
                        COUNT(*) as request_count
                    FROM help_requests
                    {date_filter}
                    GROUP BY DATE_FORMAT(created_at, '%Y-%m-01')
                    ORDER BY month_date ASC
                """)
            raw_data = cursor.fetchall()
            
            # Format dates using NZ date function
            trends = []
            for row in raw_data:
                if period in ['last_week', 'last_month']:
                    # Daily data for short periods
                    date_value = row['date']
                    if isinstance(date_value, str):
                        date_value = datetime.strptime(date_value, '%Y-%m-%d')
                    
                    nz_date_str = nz_date(date_value)  # Returns DD/MM/YYYY
                    day, month, year = nz_date_str.split('/')
                    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                    formatted_date = f"{day} {month_names[int(month)-1]}"
                    
                    trends.append({
                        'month': formatted_date,
                        'request_count': row['request_count']
                    })
                else:
                    # Monthly data for longer periods
                    month_date = row['month_date']
                    if isinstance(month_date, str):
                        month_date = datetime.strptime(month_date, '%Y-%m-%d')
                    
                    # Use nz_date and extract month/year part
                    nz_date_str = nz_date(month_date)  # Returns DD/MM/YYYY
                    # Extract month/year from DD/MM/YYYY format
                    day, month, year = nz_date_str.split('/')
                    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                    formatted_month = f"{month_names[int(month)-1]} {year}"
                    
                    trends.append({
                        'month': formatted_month,
                        'request_count': row['request_count']
                    })
            
            return convert_decimals_to_int(trends)
            
    except Exception as e:
        print(f"Error getting request trends: {e}")
        return []


def get_category_breakdown(period='all'):
    """
    Get category breakdown for help requests with percentage calculations
    """
    try:
        with db.get_cursor() as cursor:
            # Build date filter based on period
            date_filter = ""
            if period == 'last_week':
                date_filter = "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 1 WEEK)"
            elif period == 'last_month':
                date_filter = "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 1 MONTH)"
            elif period == 'last_3_months':
                date_filter = "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 3 MONTH)"
            elif period == 'last_6_months':
                date_filter = "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 6 MONTH)"
            
            # Get total count first
            cursor.execute(f"SELECT COUNT(*) as total FROM help_requests {date_filter}")
            total_result = cursor.fetchone()
            total = total_result['total'] if total_result else 0
            
            # Get category breakdown
            cursor.execute(f"""
                SELECT 
                    category,
                    COUNT(*) as count
                FROM help_requests
                {date_filter}
                GROUP BY category
                ORDER BY count DESC
            """)
            breakdown = cursor.fetchall()
            
            # Add percentage calculations
            result = []
            for item in breakdown:
                count = item['count'] or 0
                percentage = round((count / total * 100), 1) if total > 0 else 0
                result.append({
                    'category': item['category'],
                    'count': count,
                    'percentage': percentage
                })
            
            return convert_decimals_to_int(result)
            
    except Exception as e:
        print(f"Error getting category breakdown: {e}")
        return []


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def convert_decimals_to_int(data):
    # Convert Decimal to int for JSON (MySQL returns Decimal but JS needs int)
    if isinstance(data, dict):
        new_dict = {}
        for key in data:
            value = data[key]
            # Call function again for nested structures
            new_dict[key] = convert_decimals_to_int(value)
        return new_dict
    
    # Handle lists 
    elif isinstance(data, list):
        new_dict = []
        for item in data:
            new_dict.append(convert_decimals_to_int(item))
        return new_dict
    
    # Handle Decimal - checking as string cause couldn't get isinstance to work
    elif str(type(data)) == "<class 'decimal.Decimal'>":
        # For participation_rate, keep as float to preserve decimal places
        if hasattr(data, '__float__'):
            return float(data)
        return int(data)
    
    # Everything else (int, str, None, etc) just return as-is
    else: 
        return data
 
def get_personal_activity_stats(user_id):
    """
    Get personal activity statistics for a participant across all groups
    """
    try:
        with db.get_cursor() as cursor:
            # Get overall statistics
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT CASE WHEN CONCAT(e.event_date, ' ', e.event_time) < NOW() AND em.participation_status = 'attended' THEN em.event_id END) as total_attended,
                    COUNT(CASE WHEN CONCAT(e.event_date, ' ', e.event_time) < NOW() AND em.event_role = 'volunteer' AND em.participation_status = 'attended' THEN 1 END) as total_volunteer_events,
                    COALESCE(SUM(CASE WHEN CONCAT(e.event_date, ' ', e.event_time) < NOW() AND em.event_role = 'volunteer' AND em.participation_status = 'attended' THEN em.volunteer_hours ELSE 0 END), 0) as total_volunteer_hours,
                    COUNT(CASE WHEN CONCAT(e.event_date, ' ', e.event_time) > NOW() AND em.participation_status = 'registered' THEN 1 END) as upcoming_events
                FROM event_members em
                JOIN event_info e ON em.event_id = e.event_id
                WHERE em.user_id = %s
            """, (user_id,))
            overall_stats = cursor.fetchone()
            
            # Get statistics by group
            cursor.execute("""
                SELECT 
                    g.group_id,
                    g.name as group_name,
                    COUNT(DISTINCT CASE WHEN CONCAT(e.event_date, ' ', e.event_time) < NOW() AND em.participation_status = 'attended' THEN em.event_id END) as attended_count,
                    COUNT(CASE WHEN CONCAT(e.event_date, ' ', e.event_time) < NOW() AND em.event_role = 'volunteer' AND em.participation_status = 'attended' THEN 1 END) as volunteer_count,
                    COALESCE(SUM(CASE WHEN CONCAT(e.event_date, ' ', e.event_time) < NOW() AND em.event_role = 'volunteer' AND em.participation_status = 'attended' THEN em.volunteer_hours ELSE 0 END), 0) as volunteer_hours,
                    COUNT(CASE WHEN CONCAT(e.event_date, ' ', e.event_time) > NOW() AND em.participation_status = 'registered' THEN 1 END) as upcoming_count
                FROM group_members gm
                JOIN group_info g ON gm.group_id = g.group_id
                LEFT JOIN event_members em ON gm.user_id = em.user_id AND em.event_id IN (
                    SELECT event_id FROM event_info WHERE group_id = g.group_id
                )
                LEFT JOIN event_info e ON em.event_id = e.event_id
                WHERE gm.user_id = %s AND gm.status = 'active'
                GROUP BY g.group_id, g.name
                ORDER BY g.name
            """, (user_id,))
            group_stats = cursor.fetchall()
            
            # Get recent activity history (past events only) with race results for participants
            cursor.execute("""
                SELECT 
                    e.event_id,
                    e.event_title,
                    e.event_date,
                    e.event_time,
                    e.location,
                    g.name as group_name,
                    em.participation_status,
                    em.volunteer_hours,
                    em.event_role,
                    em.registration_date,
                    rr.finish_time,
                    rr.start_time,
                    rr.race_rank,
                    CASE 
                        WHEN em.event_role = 'participant' AND rr.membership_id IS NOT NULL AND rr.start_time IS NOT NULL AND rr.finish_time IS NOT NULL
                        THEN TIMESTAMPDIFF(SECOND, rr.start_time, rr.finish_time)
                        ELSE NULL
                    END as elapsed_seconds,
                    CASE 
                        WHEN em.event_role = 'participant' AND rr.membership_id IS NOT NULL THEN 1
                        WHEN em.event_role = 'volunteer' AND em.volunteer_hours IS NOT NULL THEN 1
                        ELSE 0
                    END as has_result
                FROM event_members em
                JOIN event_info e ON em.event_id = e.event_id
                JOIN group_info g ON e.group_id = g.group_id
                LEFT JOIN race_results rr ON em.membership_id = rr.membership_id
                WHERE em.user_id = %s
                  AND CONCAT(e.event_date, ' ', e.event_time) < NOW()
                ORDER BY e.event_date DESC, e.event_time DESC
            """, (user_id,))
            recent_activity = cursor.fetchall()
            
            # Convert decimals to int/float
            overall_stats = convert_decimals_to_int(overall_stats)
            group_stats = convert_decimals_to_int(group_stats)
            recent_activity = convert_decimals_to_int(recent_activity)
            
            return {
                'overall_stats': overall_stats,
                'group_stats': group_stats,
                'recent_activity': recent_activity
            }
            
    except Exception as e:
        print(f"Error getting personal activity stats for user {user_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def validate_analytics_data(data, required_keys):
    """
    Validate that analytics data contains required keys

    """
    if not data:
        return False
    
    for key in required_keys:
        if key not in data:
            print(f"Warning: Missing required key '{key}' in analytics data")
            return False
    
    return True